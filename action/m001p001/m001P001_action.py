#
# @Author: Syed Athar Hussain
# @Date: 2024-02-04
# @LastEditTime: 
# @LastEditors: 2024-03-14
# @Description: action
#
import json
import os
import openpyxl
# from common.api import langchain_api
from common.func  import common_fun
# from ...common.api import langchain_api
from common.api import langchain_api
import pandas as pd
import shutil
import re


def data_processing(url,code,languageType,content_change):
    # data = {
    #         'チケット番号': ['#3066'],
    #         '指摘内容': ['50行目付近、base.init()遺漏'],
    #         '処置内容': ['指摘内容を追加しました。'],
    #         'レビュー結果': ['NG',],
    #         '説明': ['処置内容の説明は不明確です。']
    #     }
    # df = pd.DataFrame(data)
    
    # print(df)
    #redmineからデータを取得する
    redmine_response=common_fun.get_data(url,code)
    if redmine_response[0]!=200:
        return redmine_response[1]
        # return pd.DataFrame(redmine_response[1])
    #redmineから取得したデータを処理する
    json_data=redmine_response[1]
    redmine_data =  {}
    redmine_data['チケット番号'] = []
    redmine_data['指摘内容'] = []
    redmine_data['指摘内容_api用'] = []
    redmine_data['処置内容'] = []
    redmine_data['処置内容_api用'] = []
    
    redmine_data['ステータス'] = []
    #データ整理
    for issues in json_data["issues"]:
        redmine_data['チケット番号'].append(issues["id"])
        redmine_data['ステータス'].append(issues["status"]["id"])
        for custom_field in issues["custom_fields"]:
            if custom_field["id"] ==31:
                #指摘内容
                redmine_data['指摘内容'].append(custom_field["value"])
                if "指摘内容" in custom_field["value"]:
                    redmine_data['指摘内容_api用'].append(custom_field["value"])
                else:
                    redmine_data['指摘内容_api用'].append("指摘内容："+custom_field["value"])
            # TODO 処置内容
            elif custom_field["id"] ==33 :
            # if custom_field["id"] ==33 :
                print("処置内容",custom_field)
                redmine_data['処置内容'].append(custom_field["value"])
                if "処置内容" in custom_field["value"]:
                    redmine_data['処置内容_api用'].append(custom_field["value"])
                else:
                    redmine_data['処置内容_api用'].append("処置内容："+custom_field["value"])        

    redmine_data["レビュー結果_指摘内容"]=[]
    redmine_data["レビュー結果_処置内容"]=[]
    redmine_data["不足情報_指摘内容"]=[]
    
        
    # redmine_data["説明"]=[]
    redmine_data["説明_指摘内容"]=[]
    redmine_data["説明_処置内容"]=[]
    redmine_data["不足情報_処置内容"]=[]
    
    
    #api呼び出し          
    for i in range(0,len(redmine_data["チケット番号"])): 
        
        
        # 指摘内容テスト用
        # print("嘎嘎",redmine_data["指摘内容_api用"][i])
        # result=langchain_api.langchain_api(1,redmine_data["指摘内容_api用"][i])
        
        # result1=json.loads(result)
        # return_result=result1["result"]
        # return_content=result1["content"]
        # redmine_data["レビュー結果"].append(return_result)
        # redmine_data["説明"].append(return_content)
        
        # 処置内容テスト用
        # return_result ="OK"
        # return_content =''
        # print("嘎嘎",redmine_data["処置内容"][i])
        # print("嘎嘎1",redmine_data["処置内容"][i],redmine_data["ステータス"][i])
        # if redmine_data["ステータス"][i]==5 or redmine_data["ステータス"][i]==6:
        #     if redmine_data["処置内容"][i]:
        #         print("121gag")
        #         result=langchain_api.langchain_api(3,redmine_data["処置内容_api用"][i])
        #         result1=json.loads(result)
        #         return_result=result1["result"]
        #         return_content=result1["content"]
        #         redmine_data["レビュー結果"].append(return_result)
        #         redmine_data["説明"].append(result)
        #     else:
        #         print("122gag")
        #         redmine_data["レビュー結果"].append("NG")
        #         redmine_data["説明"].append("-")
        # else:
        #     print("123gag")
        #     redmine_data["レビュー結果"].append("OK")
        #     redmine_data["説明"].append("-")
        # TODO 指摘内容と処置内容（）
        # result1=json.loads(langchain_api.langchain_api(1,redmine_data["指摘内容"][i]))
        # result2=json.loads(langchain_api.langchain_api(2,redmine_data["処置内容"][i]))
        # return_result ="OK"
        # return_content =''
        # if result1["result"]=="NG":
        #     return_result="NG"
        #     return_content=result1["content"]
        # if result2["result"]=="NG":
        #     return_result="NG"
        #     return_content+='\r\n'
        #     return_content+=result2["content"]
        # if not return_content:
        #     return_content ="-"
        # redmine_data["レビュー結果"].append(return_result)
        # redmine_data["説明"].append(return_content)
        
    
        #レビュー結果_指摘内容
        censure_result='OK'
        #説明_指摘内容 
        censure_content='' 
        
        #レビュー結果_処置内容
        process_result='OK'
        #説明_処置内容
        process_content=''
        
        
        # 指摘内容
        if content_change == "Point Out Category":
            if not (isinstance(redmine_data["指摘内容"][i], float)):
                if redmine_data["指摘内容"][i].strip():
                    
                    try:
                        result1=json.loads(langchain_api.langchain_api_review(1,redmine_data["指摘内容_api用"][i],languageType))
                        if result1["result"]=="NG":
                            censure_result="NG"
                            censure_content=result1["JP_content"]+"\r\n"+result1["EN_content"]
                    except json.JSONDecodeError as e:
                        censure_result="-"
                        censure_content="指摘内容：実行エラー"
                        
                else:
                    censure_result="-"
                    censure_content="-"

                redmine_data["レビュー結果_指摘内容"].append(censure_result)
                redmine_data["説明_指摘内容"].append(censure_content)
                redmine_data["不足情報_指摘内容"].append('')
                
                
                
        elif content_change == "Cause Classification & Category":
        # 処置内容
            if not (isinstance(redmine_data["処置内容"][i], float)):
                if redmine_data["ステータス"][i]==5 or redmine_data["ステータス"][i]==6:
                    if redmine_data["処置内容"][i].strip():
                        try:
                            result2=json.loads(langchain_api.langchain_api_review(2,redmine_data["処置内容_api用"][i],languageType))
                            if result2["result"]=="NG":
                                process_result="NG"
                                process_content=result2["JP_content"]+"\r\n"+result2["EN_content"]
                                
                                
                        except json.JSONDecodeError as e:
                            process_result="-"
                            process_content="処置内容：実行エラー"
                            
                    else:
                        process_result="-"
                        process_content="-"
                else:
                    process_result="-"
                    process_content="-"
                redmine_data["レビュー結果_処置内容"].append(process_result)
                redmine_data["説明_処置内容"].append(process_content)
                redmine_data["不足情報_処置内容"].append('')

         
             
        # redmine_data["レビュー結果_指摘内容"].append(censure_result)
        # redmine_data["説明_指摘内容"].append(censure_content)
        # redmine_data["不足情報_指摘内容"].append('')
        # # redmine_data["修正"].append(correct)
        # redmine_data["レビュー結果_処置内容"].append(process_result)
        # redmine_data["説明_処置内容"].append(process_content)
        # redmine_data["不足情報_処置内容"].append('')
        
     
        
        
    
    # del redmine_data['指摘内容_api用'] 
    # del redmine_data['処置内容_api用'] 
    # del redmine_data["ステータス"]
    # del redmine_data["説明_指摘内容"]
    # del redmine_data["説明_処置内容"]
    
 
    # df = pd.DataFrame(redmine_data)
    # df.to_excel('10.xlsx', index=False)
    return redmine_data

#補充情報の取得
def get_extra_info(data,type,url,code):
    if type==1:
        data['不足情報_指摘内容'] = []
    else:
        data['不足情報_処置内容'] = []
    
    
    
    redmine_response=common_fun.get_data(url,code)
    if redmine_response[0]!=200:
        return redmine_response
        # return pd.DataFrame(redmine_response[1])
    #redmineから取得したデータを処理する
    json_data=redmine_response[1]
    
    for i in range(0,len(data["チケット番号"])):
        #不足情報_指摘内容
        review=''
        #不足情報_処置内容
        process=''
        for issues in json_data["issues"]:
            if issues["id"]==data["チケット番号"][i]:
                for custom_field in issues["custom_fields"]:
                    #不足情報_指摘内容
                    if custom_field["id"] ==82 and type==1:
                        review=custom_field["value"]
                    #不足情報_処置内容
                    if custom_field["id"] ==8 and type!=1:
                        process=custom_field["value"]
        if type ==1:
            data['不足情報_指摘内容'].append(review)
        else:
            data['不足情報_処置内容'].append(process)
    return [200,data]


#修正   指摘内容  NG説明 補充情報
def data_revise(content,description,info,languageType):
    correct=''
    try:
        result=json.loads(langchain_api.langchain_api_revise(1,content,description,info,languageType))
        correct=result["JP_content"]+"\r\n"+result["EN_content"]
    except json.JSONDecodeError as e:
        correct="指摘内容：実行エラー" 
    return correct
#OKデータ修正   処置内容   補充情報
def ok_data_revise(content):
    correct=''
    try:
        result=json.loads(langchain_api.langchain_api_ok_revise(1,content))
        correct=result["JP_content"]+"\r\n"+result["EN_content"]
    except json.JSONDecodeError as e:
        correct="指摘内容：実行エラー" 
    return correct
 
   
#修正   処置内容  NG説明 補充情報
def data_revise_process(content,description,info,languageType):
    process=''
    try:
        result=json.loads(langchain_api.langchain_api_revise(2,content,description,info,languageType))
        process=result["JP_content"]+"\r\n"+result["EN_content"]
    except json.JSONDecodeError as e:
        process="処置内容：実行エラー" 
    return process

#OKデータ修正   処置内容   補充情報
def ok_data_revise_process(content):
    process=''
    try:
        result=json.loads(langchain_api.langchain_api_ok_revise(2,content))
        process=result["JP_content"]+"\r\n"+result["EN_content"]
    except json.JSONDecodeError as e:
        process="処置内容：実行エラー" 
    return process

#一括修正
def data_batch_revise(data,type,languageType):
    
   
    # redmine_response= get_extra_info(data,type,url,code)
    # if redmine_response[0]!=200:
    #     return redmine_response
        # return pd.DataFrame(redmine_response[1])
    #redmineから取得したデータを処理する
    json_data=data
   
    if type==1:
        json_data["修正_指摘内容"]=[] 
    else:
        json_data["修正_処置内容"]=[]
    
    print(999999,data)
    for i in range(0,len(json_data["チケット番号"])):
        if type==1:
            correct=''
            print("レビュー結果_指摘内容cesi",json_data["レビュー結果_指摘内容"][i])
            # if json_data["レビュー結果_指摘内容"][i]=='NG' and json_data["不足情報_指摘内容"][i] and type==1:
            if json_data["レビュー結果_指摘内容"][i]=='NG' :
                correct =data_revise(json_data["指摘内容"][i],json_data["説明_指摘内容"][i],json_data["不足情報_指摘内容"][i],languageType)
            elif json_data["レビュー結果_指摘内容"][i]=='OK':
                correct=ok_data_revise(json_data["指摘内容"][i])
            else:
                correct=json_data["指摘内容"][i]
            json_data["修正_指摘内容"].append(correct)
        elif type==2:
            process=''
            # if json_data["レビュー結果_処置内容"][i]=='NG' and json_data["不足情報_処置内容"][i] and type==2:
            if json_data["レビュー結果_処置内容"][i]=='NG':
                process =data_revise_process(json_data["処置内容"][i],json_data["説明_処置内容"][i],json_data["不足情報_処置内容"][i],languageType)
            elif json_data["レビュー結果_処置内容"][i]=='OK':
                process =ok_data_revise_process(json_data["処置内容"][i])
            else:
                correct=json_data["指摘内容"][i]
            json_data["修正_処置内容"].append(process) 
    return json_data

# Excelからデータを取得する
def excel_read_info(file,languageType,content_change):
    df = pd.read_excel(file, engine='openpyxl')
    print("after df in action")
    print(content_change)
    # read_file = pd.read_excel (r"./data/output/"+os.path.basename(path), sheet_name="review sheet")
    # 查找目标列的位置
   
    # # 指定要搜索的列索引（第一列索引为0）
    # column_index = 0

    # 前の16行を削除
    df1=df.drop(range(16))
    # 最初の列が空白の行を削除
    df2=df1.dropna(subset=[df1.columns[0]])
    print("df2",df2)
    # 空の列を削除
    df3=df2.dropna(axis=1, how='all')
    print("df3",df3)
    # 指摘内容が空白の行を削除
    df4=df3.dropna(subset=[df3.columns[5]])
    print("df4",df4.iloc[1:, 0])
    data={}
    data["チケット番号"]=df4.iloc[1:, 0].tolist()
    data["指摘内容"]=df4.iloc[1:, 5].tolist()
    data["処置内容"]=df4.iloc[1:, 6].tolist()
    data["レビュー結果_指摘内容"]=[]
    data["説明_指摘内容"]=[]
    data["不足情報_指摘内容"]=[]
    data["レビュー結果_処置内容"]=[]
    data["説明_処置内容"]=[]
    data["不足情報_処置内容"]=[]
    print("data",data)
    
    
    for i in range(0,len(data["チケット番号"])): 
        
    
        #レビュー結果_指摘内容
        censure_result='OK'
        #説明_指摘内容 
        censure_content='' 
        
        #レビュー結果_処置内容
        process_result='OK'
        #説明_処置内容0
        process_content=''
        
        
        # 指摘内容
        if content_change == "Point Out Category":
            if not (isinstance(data["指摘内容"][i], float)):
                if str(data["指摘内容"][i]).strip():
                    
                    try:
                        test=''
                        if "指摘内容" in data["指摘内容"][i]:
                            test=data["指摘内容"][i]
                        else:
                            test="指摘内容："+data["指摘内容"][i]
                        print("1nd loop")
                        print(i)
                        result1=json.loads(langchain_api.langchain_api_review(1,test,languageType))
                        if result1["result"]=="NG":
                            censure_result="NG"
                            censure_content=result1["JP_content"]+"\r\n"+result1["EN_content"]
                    except json.JSONDecodeError as e:
                        censure_result="-"
                        censure_content="指摘内容：実行エラー"
                        
                else:
                    censure_result="-"
                    censure_content="-"
                
                data["レビュー結果_指摘内容"].append(censure_result)
                data["説明_指摘内容"].append(censure_content)
                data["不足情報_指摘内容"].append('')
        
        elif content_change == "Cause Classification & Category":
        # 処置内容
            if not (isinstance(data["処置内容"][i], float)):
                if str(data["処置内容"][i]).strip():
                    try:
                        test1=''
                        if "処置内容" in data["処置内容"][i]:
                            test1=data["処置内容"][i]
                        else:
                            test1="処置内容"+data["処置内容"][i]
                        print("2nd loop")
                        print(i)
                        result2=json.loads(langchain_api.langchain_api_review(2,test1,languageType))
                        if result2["result"]=="NG":
                            process_result="NG"
                            process_content=result2["JP_content"]+"\r\n"+result2["EN_content"]
                            
                            
                    except json.JSONDecodeError as e:
                        process_result="-"
                        process_content="処置内容：実行エラー"
                        
                else:
                    process_result="-"
                    process_content="-"

                data["レビュー結果_処置内容"].append(process_result)
                data["説明_処置内容"].append(process_content)
                data["不足情報_処置内容"].append('')
    
    
    
    return data


def implementation_process(contents,radio_button):
    result_cause_classification = []
    result_cause_category = []
    result_point_out_category = []
    result_point_serial_no = []
    result_cause_serial_no = []
    df = pd.DataFrame()

    print("testttt--44")
    
    contents_list = contents
    categories = []
    if radio_button=="Point Out Category":
        
        categories = langchain_api.pointOutCategoryFun(contents_list)

    elif radio_button == "Cause Classification & Category":    
        categories = langchain_api.causeclassificationFun(contents_list)
        
        # print(cause_classification_category)
    #print(point_out_categories)
    if radio_button=="Point Out Category":  
         
        j=1
        k=21
        try:
            print("------------insde Review Content------------") 
            for value_cat in categories:
                regex = re.compile(r"[09]")
                if regex.search(value_cat):
                    print("inside regex-- "+value_cat)
                    if "('" in value_cat:
                        res_list = re.findall(r"('([0-9]{2}:.*?)')", value_cat, re.DOTALL)
                        if len(res_list) > 1:
                            res_list.sort()

                        listToStr = ', '.join([str(elem) for elem in res_list])
                        print("inside elif ")
                        print(" listtostr:--"+listToStr)
                        # ws["N"+str(k)] = "NA"
                        result_point_out_category.append("NA")
                    
    
                    elif "'" in value_cat:
                        res_list = re.findall(r"'([0-9]{2}:.*?)'", value_cat, re.DOTALL)
                        if len(res_list) > 1:
                            res_list.sort()

                        listToStr = ', '.join([str(elem) for elem in res_list])
                        print("listtostr:---"+listToStr)
                    
                        if not listToStr:
                            # ws["N"+str(k)] = "NA"
                            result_point_out_category.append("NA")
                        else:
                            # ws["N"+str(k)] = listToStr
                            result_point_out_category.append(res_list)
    
            
    
                    else:
                        res_list = re.findall(r"([0-9]{2}:.*?)", value_cat, re.DOTALL)
                        if len(res_list) > 1:
                            res_list.sort()

                        listToStr = ', '.join([str(elem) for elem in res_list])
                        print("in else regex--")
                        print("listtostr:--- "+listToStr)
                        if not listToStr:
                            # ws["N"+str(k)] = "NA"
                            result_point_out_category.append("NA")
                        else:
                            # ws["N"+str(k)] = listToStr
                            result_point_out_category.append(res_list)

                else:        
                    print("outside regex--")
                    print("value_cat:--- "+value_cat)
                    # ws["N"+str(k)] = "NA"
                    result_point_out_category.append("NA")
                
                # result_point_out_category.append(j)
                j=j+1
                k=k+1
        except Exception as e:
            print("Review Content Exception "+e)
        
        print("length of arrays:---")
        print(len(contents_list))
        print(len(result_point_out_category))
        result_data = [contents_list,
                   result_point_out_category]

        
        
    elif radio_button == "Cause Classification & Category":
        values = ["1:要件提示ミス", "1:要件提示ミス" , "1:要件提示ミス", "1:要件提示ミス", "2:設計ミス", "2:設計ミス",
                "2:設計ミス", "2:設計ミス", "2:設計ミス", "3:作業ミス", "3:作業ミス", "3:作業ミス", "3:作業ミス", "3:作業ミス",
                    "3:作業ミス", "4:ノウハウ不足", "5:その他", "8:設計改善", "9:指摘ミス", "9:指摘ミス"]
        keys = ['10:仕様変更（条件変更）', '11:仕様変更（仕様不具合）', '12:設計変更', '13:要件追加', 
                    '20:検討不足', '21:影響調査漏れ', '22:標準化／規約不足', '23:提示条件の確認不足', '24:関連機能の確認不足'
                    , '30:修正漏れ', '31:修正誤り', '32:標準化／規約違反', '33:周知連絡不徹底', '34:表現の配慮不足', 
                    '35:単純ミス', '40:ノウハウ不足', '50:その他', '80:設計の改善', '90:指摘ミス', '91:対応せず（仕様通り）']
          
        j = 1
        k=21
        try:
            print("------------insde Treatment Content------------")  
            for cause_cat in categories:
                regex = re.compile(r"[0-9]{2}")
                if regex.search(cause_cat):
                    print("inside regex-- "+cause_cat)
                    if "('" in cause_cat:
                        print("------------------------- :" + str(j))
                        cause_classification = []
                        cause_cat_list = re.findall(r"('([0-9]{2}:.*?)')", cause_cat, re.DOTALL)
                        for cause_cat_value in cause_cat_list:
                            cause_classification.append(values[keys.index(cause_cat_value)])

                        if len(cause_cat_list) > 0:
                            cause_cat_list.sort()
                            cause_classification.sort()

                        listToStr_cause_cat = ', '.join([str(elem) for elem in re.findall(r"('([0-9]{2}:.*?)')", cause_cat, re.DOTALL)])
                        listToStr_cause_classification = ', '.join([str(elem) for elem in cause_classification])
                        print("inside if ")
                        print(" listtostr:--"+listToStr_cause_cat)
                        print(" listtostr:--"+listToStr_cause_classification)
                        # ws["BE"+str(k)] = "NA"
                        # ws["BH"+str(k)] = "NA"
                        result_cause_classification.append("NA")
                        result_cause_category.append("NA")
                    
    
                    elif "'" in cause_cat:
                        print("------------------------- :" + str(j))
                        cause_classification = []
                        cause_cat_list = re.findall(r"'([0-9]{2}:.*?)'", cause_cat, re.DOTALL)
                        for cause_cat_value in cause_cat_list:
                            cause_classification.append(values[keys.index(cause_cat_value)])
                        
                        if len(cause_cat_list) > 0:
                            cause_cat_list.sort()
                            cause_classification.sort()
                        
                        listToStr_cause_cat = ', '.join([str(elem) for elem in re.findall(r"'([0-9]{2}:.*?)'", cause_cat, re.DOTALL)])
                        listToStr_cause_classification = ', '.join([str(elem) for elem in cause_classification])
                        print("inside elif ")
                        print(" listtostr:--"+listToStr_cause_cat)
                        print(" listtostr:--"+listToStr_cause_classification)
                    
                        if not listToStr_cause_cat:
                            # ws["BE"+str(k)] = "NA"
                            # ws["BH"+str(k)] = "NA"
                            result_cause_classification.append("NA")
                            result_cause_category.append("NA")

                        else:
                            # ws["BE"+str(k)] = listToStr_cause_classification
                            # ws["BH"+str(k)] = listToStr_cause_cat
                            result_cause_classification.append(cause_classification)
                            result_cause_category.append(cause_cat_list)
    
            
    
                    else:
                        print("------------------------- :" + str(j))
                        cause_classification = []
                        cause_cat_list = re.findall(r"([0-9]{2}:.*?)", cause_cat, re.DOTALL)
                        for cause_cat_value in cause_cat_list:
                            cause_classification.append(values[keys.index(cause_cat_value)])

                        if len(cause_cat_list) > 0:
                            cause_cat_list.sort()
                            cause_classification.sort()

                        listToStr_cause_cat = ', '.join([str(elem) for elem in re.findall(r"([0-9]{2}:.*?)", cause_cat, re.DOTALL)])
                        print("in else regex--")
                        listToStr_cause_classification = ', '.join([str(elem) for elem in cause_classification])
                        print(" listtostr:--"+listToStr_cause_cat)
                        print(" listtostr:--"+listToStr_cause_classification)

                        if not listToStr_cause_cat:
                            # ws["BE"+str(k)] = "NA"
                            # ws["BH"+str(k)] = "NA"
                            result_cause_classification.append("NA")
                            result_cause_category.append("NA")

                        else:
                            # ws["BE"+str(k)] = listToStr_cause_classification
                            # ws["BH"+str(k)] = listToStr_cause_cat
                            result_cause_classification.append(cause_classification)
                            result_cause_category.append(cause_cat_list)

                else:     
                    print("------------------------- :" + str(j))   
                    print("outside regex--")
                    print("value_cat:--- "+cause_cat)
                    # ws["BE"+str(k)] = "NA"
                    # ws["BH"+str(k)] = "NA"
                    result_cause_classification.append("NA")
                    result_cause_category.append("NA")
                
                result_cause_serial_no.append(j)
                j=j+1
                k=k+1

        except Exception as e:

            print("treatment content Exception "+e)
        
        print("length of arrays:---")
        print(len(contents_list))
        print(len(result_cause_classification))
        print(len(result_cause_category))

        result_data = [contents_list
                   , result_cause_classification,result_cause_category]

    
        
    print("Finished")        
    return result_data


if __name__ == "__main__":
    url="http://10.167.84.241/redmine/projects/dydo/issues.json?tracker_id=32"
    code="3963"
    languageType="1"
    data_processing(url,code,languageType)