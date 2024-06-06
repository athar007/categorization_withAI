#
# @Author: Syed Athar Hussain
# @Date: 2024-02-04
# @LastEditTime: 
# @LastEditors: 2024-03-14
# @Description:指摘票レビュー画面(Redmine)
#
import os
import shutil
import gradio as gr
import pandas as pd
from openpyxl import load_workbook
from ..assets.const.m001P001_const import (CUSTOM_CSS,REDMINE_URL,PARENT_NUMBER,REVIEW_LABLE,CLEAR_LABLE,
                                           REVIEW_LIST,REVIEW_CONTENT,TREATMENT_CONTENT,CORRECTION,
                                           CLASSIFICATION,EXCEL_PATH,TITLE,ISSUE_NO,RESULT,COMMENT,
                                           INSUFFICIENT_INFOR,REVISED_ARTICLE,REFERENCE, POINT_CATEGORY, FINAL_CATEGORY, CAUSE_CATEGORY, CAUSE_CLASSIFIATION)
from action.m001p001.m001P001_action import (data_processing,get_extra_info,data_batch_revise,excel_read_info,implementation_process)


# import web as const
class m001P001_view:
    def __init__(self):
        # 言語フラグ
        self.language= 2
        # redmine or excel
        self.input_type= 1
        
        # redmineモード
        # データ
        self.global_data= {}   
        # redmine_url     
        self.redmine_url=None
        # 親チケット番号
        self.parent_no=None
        # レビュー
        self.review_redmine=None
        # クリア
        self.clear_redmine=None
        self.content_select=None

        self.excel_file_download=None
        # 指摘内容　修正
        self.modify1_redmine=None
        # 処置内容　修正
        self.modify2_redmine=None
        # 指摘内容　区分
        self.sort1_redmine=None
        # 処置内容　区分
        self.sort2_redmine=None
        # 指摘内容　結果
        self.global_correction=None
        # 処置内容　結果
        self.global_table=None

        self.category_generate=None

        self.extract_to_excel_file=None
        
        # excel モード
        # データ
        #self.excel_data = {}
        # excelファイルパス
        self.excel_path=None 
        # excelファイルアップロード
        self.file_upload=None
        # レビュー   
        self.review_excel=None
        # クリア
        self.clear_excel=None
        # 指摘内容　修正
        self.modify1_excel=None
        # 処置内容　修正
        self.modify2_excel=None
        # 指摘内容　区分
        self.sort1_excel=None
        # 処置内容　区分
        self.sort2_excel=None
        # 指摘内容　結果
        self.excel_result1=None
        # 処置内容　結果
        self.excel_result2=None

    # redmine レビュー処理
    def redmine_review(self,a,b,content_change):
        redmine_data=data_processing(a,b,self.language,content_change)
        print("inside redmine review")
        print(redmine_data)
        self.global_data=dict(redmine_data)
        if(ISSUE_NO[1] in redmine_data):
            print(content_change)
            
            if content_change=="Point Out Category":
                data1={}
                data1[ISSUE_NO[self.language]]  = redmine_data[ISSUE_NO[1]]
                data1[REVIEW_CONTENT[self.language]]  = redmine_data[REVIEW_CONTENT[1]]
                data1[RESULT[self.language]]  = redmine_data["レビュー結果_指摘内容"]
                data1[COMMENT[self.language]]  = redmine_data["説明_指摘内容"]
                data1[INSUFFICIENT_INFOR[self.language]]  =  redmine_data["不足情報_指摘内容"]
                if not (isinstance(redmine_data[REVIEW_CONTENT[1]][0], float)):
                    return data1
                else:
                    return {}
            
            elif content_change == "Cause Classification & Category":
                data2={}
                data2[ISSUE_NO[self.language]]  = redmine_data[ISSUE_NO[1]]
                data2[TREATMENT_CONTENT[self.language]]  = redmine_data[TREATMENT_CONTENT[1]]
                data2[RESULT[self.language]]  = redmine_data["レビュー結果_処置内容"]
                data2[COMMENT[self.language]]  = redmine_data["説明_処置内容"]
                data2[INSUFFICIENT_INFOR[self.language]]  =  redmine_data["不足情報_処置内容"]
                if not (isinstance(redmine_data[TREATMENT_CONTENT[1]][0], float)):
                    return data2
                else:
                    return {}
            
        if content_change=="Point Out Category":
            return redmine_data
        elif content_change == "Cause Classification & Category":
            return redmine_data
    
    # excel レビュー処理
    def excel_review(self,path,content_change):
        print("in excel folder")
        print(path)
        data=excel_read_info(path,self.language,content_change)
        print("Excel Data")
        print(data)
        self.global_data=dict(data)
        if(ISSUE_NO[1] in data):
            if content_change=="Point Out Category":
                data1={}
                data1[ISSUE_NO[self.language]]  = data[ISSUE_NO[1]]
                data1[REVIEW_CONTENT[self.language]]  = data[REVIEW_CONTENT[1]]
                data1[RESULT[self.language]]  = data["レビュー結果_指摘内容"]
                data1[COMMENT[self.language]]  = data["説明_指摘内容"]
                data1[INSUFFICIENT_INFOR[self.language]]  =  data["不足情報_指摘内容"]
                if not (isinstance(data[REVIEW_CONTENT[1]][0], float)):
                    return data1
                else:
                    return {}
            
            elif content_change == "Cause Classification & Category":
                data2={}
                data2[ISSUE_NO[self.language]]  = data[ISSUE_NO[1]]
                data2[TREATMENT_CONTENT[self.language]]  = data[TREATMENT_CONTENT[1]]
                data2[RESULT[self.language]]  = data["レビュー結果_処置内容"]
                data2[COMMENT[self.language]]  = data["説明_処置内容"]
                data2[INSUFFICIENT_INFOR[self.language]]  =  data["不足情報_処置内容"]
                if not (isinstance(data[TREATMENT_CONTENT[1]][0], float)):
                    return data2
                else:
                    return {}
            
        if content_change=="Point Out Category":
            return data
        elif content_change == "Cause Classification & Category":
            return data              
    
    def global_modify(self,viewData,content_select):
        data = viewData
        print("global data")
        print(self.global_data)
        if content_select == "Point Out Category":
            for i in range(0,len(data[ISSUE_NO[self.language]])):
                for j in range(0,len(self.global_data["チケット番号"])):
                    if data[ISSUE_NO[self.language]][i]==self.global_data[ISSUE_NO[1]][j]:
                        self.global_data["不足情報_指摘内容"][j]=data[INSUFFICIENT_INFOR[self.language]][i]
            excel_data=data_batch_revise(self.global_data,1,self.language)
            self.global_data=dict(excel_data)
            if("チケット番号" in excel_data):
            
                data1={}
                data1[ISSUE_NO[self.language]]  = excel_data[ISSUE_NO[1]] 
                data1[REVIEW_CONTENT[self.language]]  = excel_data[REVIEW_CONTENT[1]]
                data1[INSUFFICIENT_INFOR[self.language]]  = excel_data["不足情報_指摘内容"]
                data1[REVISED_ARTICLE[self.language]]  = excel_data["修正_指摘内容"]
                return data1
            return excel_data
        else:
            for i in range(0,len(data[ISSUE_NO[self.language]])):
                for j in range(0,len(self.global_data["チケット番号"])):
                    if data[ISSUE_NO[self.language]][i]==self.global_data[ISSUE_NO[1]][j]:
                        self.global_data["不足情報_処置内容"][j]=data[INSUFFICIENT_INFOR[self.language]][i]
            excel_data=data_batch_revise(self.global_data,2,self.language)
            self.global_data=dict(excel_data)
            if("チケット番号" in excel_data):
            
                data1={}
                data1[ISSUE_NO[self.language]]  = excel_data[ISSUE_NO[1]] 
                data1[TREATMENT_CONTENT[self.language]]  = excel_data[TREATMENT_CONTENT[1]]
                data1[INSUFFICIENT_INFOR[self.language]]  = excel_data["不足情報_処置内容"]
                data1[REVISED_ARTICLE[self.language]]  = excel_data["修正_処置内容"]
                print(data1)
                return data1
            return excel_data
    
    
    def category_suggestion(self, viewData, content_select):
        contents=[]
        data=viewData
        for value in data["Revised article"]:
            contents.append(value)

        if not contents:
            print("not")
            return  {}

        elif contents:
            
            result_contents = implementation_process(contents,content_select)
            try:
                print("inside try block")
                if content_select=="Point Out Category":
                    data1={}
                    data1[ISSUE_NO[self.language]]  = data[ISSUE_NO[2]] 
                    data1[REVISED_ARTICLE[self.language]]  = data[REVISED_ARTICLE[2]]
                    data1[POINT_CATEGORY[self.language]]  = result_contents[1]
                    data1[FINAL_CATEGORY[self.language]]  = []
                    return data1

                else:
                    data2={}
                    data2[ISSUE_NO[self.language]]  = data[ISSUE_NO[2]] 
                    data2[REVISED_ARTICLE[self.language]]  = data[REVISED_ARTICLE[2]]
                    data2[CAUSE_CATEGORY[self.language]]  = result_contents[2]
                    data2[CAUSE_CLASSIFIATION[self.language]]  = result_contents[1]
                    data2[FINAL_CATEGORY[self.language]]  = []
                    return data2
                    
            except Exception as e:
                print("inside except block")
                print(e)
                return {}
                
    def process_table_to_excel(self, viewdata, content_cat):
        file = "./data/input/00_review-sheet.xlsm"
        result_path = "./data/output/output_00_review-sheet.xlsm"
        shutil.copyfile(file, result_path)
        print("file copied")
        try:
            wb = load_workbook(filename=result_path, read_only=False , keep_vba=True)
            ws = wb['review sheet']
            k=21
            if content_cat =="Point Out Category":
                for value in viewdata['Final Category']:
                    ws["N"+str(k)] = value
                    k=k+1

                k=21
                for data in viewdata['Revised article']:
                    ws["R"+str(k)] = data
                    k=k+1

            elif content_cat == "Cause Classification & Category":
                suggestion = ["1:要件提示ミス", "1:要件提示ミス" , "1:要件提示ミス", "1:要件提示ミス", "2:設計ミス", "2:設計ミス",
                "2:設計ミス", "2:設計ミス", "2:設計ミス", "3:作業ミス", "3:作業ミス", "3:作業ミス", "3:作業ミス", "3:作業ミス",
                    "3:作業ミス", "4:ノウハウ不足", "5:その他", "8:設計改善", "9:指摘ミス", "9:指摘ミス"]
                category = ['10:仕様変更（条件変更）', '11:仕様変更（仕様不具合）', '12:設計変更', '13:要件追加', 
                    '20:検討不足', '21:影響調査漏れ', '22:標準化／規約不足', '23:提示条件の確認不足', '24:関連機能の確認不足'
                    , '30:修正漏れ', '31:修正誤り', '32:標準化／規約違反', '33:周知連絡不徹底', '34:表現の配慮不足', 
                    '35:単純ミス', '40:ノウハウ不足', '50:その他', '80:設計の改善', '90:指摘ミス', '91:対応せず（仕様通り）']
                for value in viewdata['Final Category']:
                    i = category.index(value)
                    category_suggestion = suggestion[i]
                    ws["BE"+str(k)] = category_suggestion
                    k=k+1

                k=21
                for data in viewdata['Final Category']:
                    ws["BH"+str(k)] = data
                    k=k+1

                k=21
                for data_desc in viewdata['Revised article']:
                    ws["AJ"+str(k)] = data_desc
                    k=k+1

            wb.save(result_path)
            return [1,result_path]

        except Exception as e:
            print("treatment content Exception "+e)
            return [1, e]